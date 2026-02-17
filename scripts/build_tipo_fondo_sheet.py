#!/usr/bin/env python3
"""
Build or replace a normalized 'Tipo_Fondo' worksheet in an Excel workbook.

Output sheet columns:
- Company_ID
- Tipo_Fondo

One Company_ID can map to multiple Tipo_Fondo rows.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl


COMPANY_COL_CANDIDATES = ("Company_ID", "CompanyID", "Company Id")
TIPO_COL_CANDIDATES = ("Tipo Fondo", "Tipo_Fondo", "TipoFondo")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u00a0", " ")
    return " ".join(text.strip().split())


def split_tipo_fondo(raw: str) -> list[str]:
    """
    Split common multi-value separators while avoiding over-splitting words.
    """
    if not raw:
        return []
    parts = re.split(r"[\r\n,;/|]+", raw)
    out: list[str] = []
    for part in parts:
        item = clean_text(part)
        if item:
            out.append(item)
    return out


def find_column_index(header_row: list[object], candidates: tuple[str, ...]) -> int:
    normalized = {clean_text(v).lower(): i for i, v in enumerate(header_row, start=1)}
    for candidate in candidates:
        idx = normalized.get(candidate.lower())
        if idx:
            return idx
    raise ValueError(f"Could not find any column from: {candidates}")


def build_tipo_fondo_sheet(
    workbook_path: Path,
    source_sheet: str = "Inversion",
    output_sheet: str = "Tipo_Fondo",
) -> tuple[int, int]:
    wb = openpyxl.load_workbook(workbook_path)
    if source_sheet not in wb.sheetnames:
        raise ValueError(f"Source sheet '{source_sheet}' not found. Sheets: {wb.sheetnames}")

    src = wb[source_sheet]
    headers = [cell.value for cell in src[1]]
    company_col = find_column_index(headers, COMPANY_COL_CANDIDATES)
    tipo_col = find_column_index(headers, TIPO_COL_CANDIDATES)

    if output_sheet in wb.sheetnames:
        del wb[output_sheet]
    dest = wb.create_sheet(output_sheet)
    dest.cell(row=1, column=1, value="Company_ID")
    dest.cell(row=1, column=2, value="Tipo_Fondo")

    seen: set[tuple[str, str]] = set()
    written = 0
    for row in range(2, src.max_row + 1):
        company_id = clean_text(src.cell(row=row, column=company_col).value)
        raw_tipo = clean_text(src.cell(row=row, column=tipo_col).value)
        if not company_id or not raw_tipo:
            continue

        for tipo in split_tipo_fondo(raw_tipo):
            key = (company_id, tipo)
            if key in seen:
                continue
            seen.add(key)
            written += 1
            dest.cell(row=written + 1, column=1, value=company_id)
            dest.cell(row=written + 1, column=2, value=tipo)

    dest.column_dimensions["A"].width = 14
    dest.column_dimensions["B"].width = 34

    wb.save(workbook_path)
    return written, len(seen)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create/replace a normalized 'Tipo_Fondo' sheet from source data."
    )
    parser.add_argument("workbook", type=Path, help="Path to .xlsx workbook")
    parser.add_argument(
        "--source-sheet",
        default="Inversion",
        help="Source worksheet name (default: Inversion)",
    )
    parser.add_argument(
        "--output-sheet",
        default="Tipo_Fondo",
        help="Output worksheet name (default: Tipo_Fondo)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1
    if workbook_path.suffix.lower() != ".xlsx":
        print("Only .xlsx files are supported.", file=sys.stderr)
        return 1

    try:
        written, unique_pairs = build_tipo_fondo_sheet(
            workbook_path=workbook_path,
            source_sheet=args.source_sheet,
            output_sheet=args.output_sheet,
        )
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print(
        f"Done. Sheet '{args.output_sheet}' written in '{workbook_path.name}'. "
        f"Rows: {written}, unique pairs: {unique_pairs}."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
