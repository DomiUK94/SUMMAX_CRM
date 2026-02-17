#!/usr/bin/env python3
"""
Compliant enrichment for investor/contact workbooks.

What this script does:
- Reads companies from the 'Inversion' sheet.
- Scrapes only public company websites (never scrapes LinkedIn pages).
- Respects robots.txt rules and applies request throttling.
- Extracts basic public signals: emails, phones, and LinkedIn company/profile links
  found on allowed website pages.
- Updates 'Inversion' and 'Contactos' sheets.
- Highlights all created/changed cells in yellow for manual validation.

Usage:
  python scripts/enrich_investors_contacts.py "260216 Mapping SUMMAX Reestructura 1.xlsx"
  python scripts/enrich_investors_contacts.py "file.xlsx" --max-companies 40 --delay 2.0
  python scripts/enrich_investors_contacts.py "file.xlsx" --dry-run
"""

from __future__ import annotations

import argparse
import datetime as dt
import html
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib import error as urlerror
from urllib import parse as urlparse
from urllib import request as urlrequest
from urllib.robotparser import RobotFileParser

import openpyxl
from openpyxl.styles import PatternFill


YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")
USER_AGENT = "SUMMAX-CRM-Enricher/1.0 (+public-site-only; contact: internal)"

EMAIL_RE = re.compile(r"\b[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[A-Za-z]{2,}\b")
PHONE_RE = re.compile(r"(?:(?:\+\d{1,3}[\s\-]?)?(?:\(?\d{2,4}\)?[\s\-]?)?\d[\d\s\-]{6,}\d)")
HREF_RE = re.compile(
    r"""href\s*=\s*["']([^"'#]+)["']""",
    flags=re.IGNORECASE,
)

CONTACT_HINTS = ("contact", "contacto", "about", "team", "equipo")


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\u00a0", " ")
    return " ".join(text.split())


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").strip().split())


def canonicalize_url(raw: str) -> Optional[str]:
    if not raw:
        return None
    raw = raw.strip()
    if not raw:
        return None
    if not re.match(r"^https?://", raw, flags=re.IGNORECASE):
        raw = "https://" + raw
    parsed = urlparse.urlparse(raw)
    if not parsed.netloc:
        return None
    return urlparse.urlunparse(
        (
            parsed.scheme.lower(),
            parsed.netloc.lower(),
            parsed.path or "/",
            "",
            "",
            "",
        )
    )


def same_domain(url: str, base: str) -> bool:
    u = urlparse.urlparse(url).netloc.lower()
    b = urlparse.urlparse(base).netloc.lower()
    return u == b


def extract_links(base_url: str, html_text: str) -> List[str]:
    links: List[str] = []
    for href in HREF_RE.findall(html_text):
        full = urlparse.urljoin(base_url, html.unescape(href.strip()))
        parsed = urlparse.urlparse(full)
        if parsed.scheme not in ("http", "https"):
            continue
        links.append(
            urlparse.urlunparse((parsed.scheme, parsed.netloc, parsed.path or "/", "", "", ""))
        )
    return links


def request_text(url: str, timeout: float = 10.0) -> Optional[str]:
    req = urlrequest.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urlrequest.urlopen(req, timeout=timeout) as resp:
            ctype = resp.headers.get("Content-Type", "").lower()
            if "text/html" not in ctype and "text/plain" not in ctype:
                return None
            raw = resp.read()
            # Best-effort decode
            enc = "utf-8"
            if "charset=" in ctype:
                enc = ctype.split("charset=", 1)[1].split(";")[0].strip() or "utf-8"
            try:
                return raw.decode(enc, errors="ignore")
            except LookupError:
                return raw.decode("utf-8", errors="ignore")
    except (urlerror.URLError, TimeoutError, ValueError):
        return None


@dataclass
class ScrapeResult:
    source_url: str
    scanned_pages: int
    emails: Set[str]
    phones: Set[str]
    linkedin_urls: Set[str]


class PoliteScraper:
    def __init__(self, delay_seconds: float):
        self.delay_seconds = delay_seconds
        self.robot_cache: Dict[str, RobotFileParser] = {}
        self.last_request_at: Dict[str, float] = {}

    def _robot_for(self, base_url: str) -> RobotFileParser:
        host = urlparse.urlparse(base_url).netloc.lower()
        if host in self.robot_cache:
            return self.robot_cache[host]
        rp = RobotFileParser()
        rp.set_url(f"https://{host}/robots.txt")
        try:
            rp.read()
        except Exception:
            # If robots is unavailable, default to permissive behavior.
            pass
        self.robot_cache[host] = rp
        return rp

    def _allowed(self, url: str) -> bool:
        host = urlparse.urlparse(url).netloc.lower()
        rp = self._robot_for(url)
        try:
            return rp.can_fetch(USER_AGENT, url) and rp.can_fetch("*", url)
        except Exception:
            return True

    def _throttle(self, url: str) -> None:
        host = urlparse.urlparse(url).netloc.lower()
        now = time.time()
        last = self.last_request_at.get(host, 0.0)
        wait = self.delay_seconds - (now - last)
        if wait > 0:
            time.sleep(wait)
        self.last_request_at[host] = time.time()

    def scrape_website(self, base_url: str, max_pages: int = 4) -> Optional[ScrapeResult]:
        start_url = canonicalize_url(base_url)
        if not start_url:
            return None
        if not self._allowed(start_url):
            return None

        queue: List[str] = [start_url]
        visited: Set[str] = set()
        emails: Set[str] = set()
        phones: Set[str] = set()
        linkedin_urls: Set[str] = set()

        while queue and len(visited) < max_pages:
            url = queue.pop(0)
            if url in visited:
                continue
            if not self._allowed(url):
                continue
            self._throttle(url)
            text = request_text(url)
            if not text:
                visited.add(url)
                continue
            visited.add(url)

            for email in EMAIL_RE.findall(text):
                if "example." in email.lower():
                    continue
                emails.add(email.lower())

            for ph in PHONE_RE.findall(text):
                # Lightweight cleanup; keep original for manual review.
                ph_clean = " ".join(ph.split())
                if len(re.sub(r"\D", "", ph_clean)) >= 8:
                    phones.add(ph_clean)

            links = extract_links(url, text)
            for link in links:
                lnk = link.lower()
                if "linkedin.com/" in lnk:
                    linkedin_urls.add(link)
                    continue
                if same_domain(link, start_url):
                    path = urlparse.urlparse(link).path.lower()
                    if any(hint in path for hint in CONTACT_HINTS):
                        queue.append(link)

        return ScrapeResult(
            source_url=start_url,
            scanned_pages=len(visited),
            emails=emails,
            phones=phones,
            linkedin_urls=linkedin_urls,
        )


def header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_header(ws.cell(1, col).value)
        if key and key not in out:
            out[key] = col
    return out


def ensure_column(
    ws: openpyxl.worksheet.worksheet.Worksheet, header: str, aliases: Sequence[str]
) -> int:
    hm = header_map(ws)
    for alias in aliases:
        idx = hm.get(normalize_header(alias))
        if idx:
            return idx
    idx = ws.max_column + 1
    ws.cell(1, idx, header)
    return idx


def set_with_highlight(
    ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int, value: str
) -> bool:
    current = clean_text(ws.cell(row, col).value)
    if current == clean_text(value):
        return False
    ws.cell(row, col, value)
    ws.cell(row, col).fill = YELLOW_FILL
    return True


def append_contact_row(
    ws_contactos: openpyxl.worksheet.worksheet.Worksheet,
    col: Dict[str, int],
    company: str,
    email: Optional[str],
    phone: Optional[str],
    linkedin: Optional[str],
    source_url: str,
) -> int:
    row = ws_contactos.max_row + 1
    ws_contactos.cell(row, col["compania"], company)
    ws_contactos.cell(row, col["persona"], "")
    ws_contactos.cell(row, col["rol"], "Website lead")
    ws_contactos.cell(row, col["otro"], "")
    ws_contactos.cell(row, col["telefono"], phone or "")
    ws_contactos.cell(row, col["email"], email or "")
    ws_contactos.cell(row, col["linkedin"], linkedin or "")
    ws_contactos.cell(row, col["comentarios"], f"Fuente web: {source_url}")
    ws_contactos.cell(row, col["prioritario"], "")

    for c in col.values():
        ws_contactos.cell(row, c).fill = YELLOW_FILL
    return row


def build_contact_unique_key(
    company: str, email: Optional[str], phone: Optional[str], linkedin: Optional[str]
) -> Tuple[str, str, str, str]:
    return (
        clean_text(company).lower(),
        clean_text(email).lower(),
        clean_text(phone).lower(),
        clean_text(linkedin).lower(),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Safely enrich Inversion/Contactos from public websites (LinkedIn-safe mode)."
    )
    parser.add_argument("workbook", type=Path, help="Path to workbook (.xlsx)")
    parser.add_argument("--max-companies", type=int, default=9999, help="Max companies to process")
    parser.add_argument("--max-pages", type=int, default=4, help="Max pages per company website")
    parser.add_argument("--delay", type=float, default=1.5, help="Delay between requests per host")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Run extraction only; do not save workbook",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    wb_path = args.workbook.expanduser().resolve()
    if not wb_path.exists():
        print(f"Workbook not found: {wb_path}", file=sys.stderr)
        return 1
    if wb_path.suffix.lower() != ".xlsx":
        print("Only .xlsx is supported", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(wb_path)
    if "Inversion" not in wb.sheetnames or "Contactos" not in wb.sheetnames:
        print("Workbook must contain sheets 'Inversion' and 'Contactos'", file=sys.stderr)
        return 1

    ws_inv = wb["Inversion"]
    ws_con = wb["Contactos"]

    inv_company_id = ensure_column(ws_inv, "Company_ID", ("Company_ID", "CompanyID"))
    inv_company = ensure_column(ws_inv, "Compañia", ("Compañia", "Compania", "Company"))
    inv_web = ensure_column(ws_inv, "Web", ("Web", "Web ", "Website"))
    inv_linkedin = ensure_column(ws_inv, "LinkedIn", ("LinkedIn", "LinkedIn ", "Linkedin"))
    inv_comments = ensure_column(ws_inv, "Comentarios", ("Comentarios",))

    con_cols = {
        "compania": ensure_column(ws_con, "Compañia", ("Compañia", "Compania", "Company")),
        "persona": ensure_column(ws_con, "Persona de contacto", ("Persona de contacto",)),
        "rol": ensure_column(ws_con, "Rol", ("Rol",)),
        "otro": ensure_column(ws_con, "Otro contacto", ("Otro contacto",)),
        "telefono": ensure_column(ws_con, "Teléfono", ("Teléfono", "Telefono", "Phone")),
        "email": ensure_column(ws_con, "email", ("email", "Email")),
        "linkedin": ensure_column(ws_con, "LinkedIn", ("LinkedIn", "Linkedin")),
        "comentarios": ensure_column(ws_con, "Comentarios", ("Comentarios",)),
        "prioritario": ensure_column(ws_con, "Prioritario?", ("Prioritario?",)),
    }

    # Build existing contact keys to avoid duplicates.
    existing_contacts: Set[Tuple[str, str, str, str]] = set()
    for r in range(2, ws_con.max_row + 1):
        key = build_contact_unique_key(
            ws_con.cell(r, con_cols["compania"]).value,
            ws_con.cell(r, con_cols["email"]).value,
            ws_con.cell(r, con_cols["telefono"]).value,
            ws_con.cell(r, con_cols["linkedin"]).value,
        )
        existing_contacts.add(key)

    scraper = PoliteScraper(delay_seconds=args.delay)
    changed_cells = 0
    created_contacts = 0
    processed = 0
    skipped_no_web = 0
    scraped_ok = 0
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    for row in range(2, ws_inv.max_row + 1):
        if processed >= args.max_companies:
            break

        company = clean_text(ws_inv.cell(row, inv_company).value)
        website = clean_text(ws_inv.cell(row, inv_web).value)
        if not company:
            continue
        processed += 1

        if not website:
            skipped_no_web += 1
            continue

        result = scraper.scrape_website(website, max_pages=args.max_pages)
        if not result:
            continue
        scraped_ok += 1

        # Inversion updates (website normalized, one linkedin if empty, comment append).
        if set_with_highlight(ws_inv, row, inv_web, result.source_url):
            changed_cells += 1

        existing_li = clean_text(ws_inv.cell(row, inv_linkedin).value)
        if not existing_li and result.linkedin_urls:
            first_li = sorted(result.linkedin_urls)[0]
            if set_with_highlight(ws_inv, row, inv_linkedin, first_li):
                changed_cells += 1

        notes = []
        if result.emails:
            notes.append(f"emails={', '.join(sorted(result.emails)[:3])}")
        if result.phones:
            notes.append(f"phones={', '.join(sorted(result.phones)[:2])}")
        if result.linkedin_urls:
            notes.append(f"linkedin={len(result.linkedin_urls)} link(s)")
        if notes:
            base = clean_text(ws_inv.cell(row, inv_comments).value)
            addon = f"[Web enrich {now}] {result.source_url} ({'; '.join(notes)})"
            new_val = (base + " | " + addon).strip(" |") if base else addon
            if set_with_highlight(ws_inv, row, inv_comments, new_val):
                changed_cells += 1

        # Contactos rows from discovered signals.
        candidate_email = sorted(result.emails)[0] if result.emails else None
        candidate_phone = sorted(result.phones)[0] if result.phones else None
        candidate_li = sorted(result.linkedin_urls)[0] if result.linkedin_urls else None

        key = build_contact_unique_key(company, candidate_email, candidate_phone, candidate_li)
        if (candidate_email or candidate_phone or candidate_li) and key not in existing_contacts:
            append_contact_row(
                ws_contactos=ws_con,
                col=con_cols,
                company=company,
                email=candidate_email,
                phone=candidate_phone,
                linkedin=candidate_li,
                source_url=result.source_url,
            )
            existing_contacts.add(key)
            created_contacts += 1

    if args.dry_run:
        print("Dry-run complete; workbook not saved.")
    else:
        wb.save(wb_path)
        print(f"Saved workbook: {wb_path.name}")

    print(
        "Summary: "
        f"processed={processed}, scraped_ok={scraped_ok}, skipped_no_web={skipped_no_web}, "
        f"changed_cells={changed_cells}, created_contact_rows={created_contacts}, dry_run={args.dry_run}"
    )
    print("LinkedIn note: this script does not scrape linkedin.com pages directly.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
