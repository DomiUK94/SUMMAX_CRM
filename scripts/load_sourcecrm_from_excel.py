#!/usr/bin/env python3
"""
Build SQL upsert script for sourcecrm tables from 260217 Mapping SUMMAX.xlsx.

Sheets supported:
- Inversion
- Contactos
- Tipo_Fondo
- Sector
- Mercados_Area_Geografica (mapped into sourcecrm.mapa_area_geografica)

The script does not connect to DB; it generates deterministic SQL ready to run.
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").strip().split())


def norm(value: object) -> str:
    text = clean(value).lower()
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def sql_quote(value: Optional[str]) -> str:
    if value is None or value == "":
        return "null"
    return "'" + value.replace("'", "''") + "'"


def parse_int(value: object) -> Optional[int]:
    raw = clean(value)
    if not raw:
        return None
    try:
        return int(raw)
    except ValueError:
        return None


def header_index(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = norm(ws.cell(1, c).value)
        if key and key not in out:
            out[key] = c
    return out


def resolve_col(index: Dict[str, int], aliases: Sequence[str], sheet_name: str) -> int:
    for alias in aliases:
        col = index.get(norm(alias))
        if col:
            return col
    raise ValueError(f"Missing required column in '{sheet_name}'. Expected one of: {aliases}")


def read_excel_rows(path: Path) -> Dict[str, List[Tuple]]:
    wb = openpyxl.load_workbook(path, data_only=True)

    required_sheets = ["Inversion", "Contactos", "Tipo_Fondo", "Sector", "Mercados_Area_Geografica"]
    for sheet in required_sheets:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Missing required sheet: {sheet}")

    inversion_rows: List[Tuple] = []
    contactos_rows: List[Tuple] = []
    tipo_rows: List[Tuple] = []
    sector_rows: List[Tuple] = []
    mapa_rows: List[Tuple] = []

    ws = wb["Inversion"]
    idx = header_index(ws)
    c_company = resolve_col(idx, ["Company_ID", "CompanyID"], "Inversion")
    c_vertical = resolve_col(idx, ["Vertical"], "Inversion")
    c_compania = resolve_col(idx, ["Compañia", "Compania"], "Inversion")
    c_direccion = resolve_col(idx, ["Dirección", "Direccion"], "Inversion")
    c_estrategia = resolve_col(idx, ["Estrategia"], "Inversion")
    c_linkedin = resolve_col(idx, ["LinkedIn", "LinkedIn "], "Inversion")
    c_web = resolve_col(idx, ["Web", "Web "], "Inversion")
    c_portfolio = resolve_col(idx, ["Portfolio"], "Inversion")
    c_comentarios = resolve_col(idx, ["Comentarios"], "Inversion")
    c_encaje = resolve_col(idx, ["Encaje SUMMAX"], "Inversion")
    c_motivo = resolve_col(idx, ["Motivo"], "Inversion")
    c_inv_min = resolve_col(idx, ["Inversion Minima"], "Inversion")
    c_inv_max = resolve_col(idx, ["Inversion Maxima"], "Inversion")
    c_prioridad = resolve_col(idx, ["Prioridad"], "Inversion")
    c_sede = resolve_col(idx, ["Sede"], "Inversion")
    c_size = resolve_col(idx, ["Tamaño Empresa", "Tamano Empresa"], "Inversion")

    for r in range(2, ws.max_row + 1):
        company_id = parse_int(ws.cell(r, c_company).value)
        if company_id is None:
            continue
        inversion_rows.append(
            (
                company_id,
                clean(ws.cell(r, c_vertical).value),
                clean(ws.cell(r, c_compania).value),
                clean(ws.cell(r, c_direccion).value),
                clean(ws.cell(r, c_estrategia).value),
                clean(ws.cell(r, c_linkedin).value),
                clean(ws.cell(r, c_web).value),
                clean(ws.cell(r, c_portfolio).value),
                clean(ws.cell(r, c_comentarios).value),
                clean(ws.cell(r, c_encaje).value),
                clean(ws.cell(r, c_motivo).value),
                clean(ws.cell(r, c_inv_min).value),
                clean(ws.cell(r, c_inv_max).value),
                clean(ws.cell(r, c_prioridad).value),
                clean(ws.cell(r, c_sede).value),
                clean(ws.cell(r, c_size).value),
            )
        )

    ws = wb["Contactos"]
    idx = header_index(ws)
    c_company = resolve_col(idx, ["Company_ID", "CompanyID"], "Contactos")
    c_contact = resolve_col(idx, ["Contact_ID", "ContactID"], "Contactos")
    c_compania = resolve_col(idx, ["Compañia", "Compania"], "Contactos")
    c_persona = resolve_col(idx, ["Persona de contacto"], "Contactos")
    c_rol = resolve_col(idx, ["Rol"], "Contactos")
    c_otro = resolve_col(idx, ["Otro contacto"], "Contactos")
    c_tel = resolve_col(idx, ["Teléfono", "Telefono"], "Contactos")
    c_email = resolve_col(idx, ["email", "Email"], "Contactos")
    c_linkedin = resolve_col(idx, ["LinkedIn", "Linkedin"], "Contactos")
    c_comentarios = resolve_col(idx, ["Comentarios"], "Contactos")
    c_prioritario = resolve_col(idx, ["Prioritario?"], "Contactos")

    for r in range(2, ws.max_row + 1):
        company_id = parse_int(ws.cell(r, c_company).value)
        contact_id = parse_int(ws.cell(r, c_contact).value)
        if company_id is None or contact_id is None:
            continue
        contactos_rows.append(
            (
                contact_id,
                company_id,
                clean(ws.cell(r, c_compania).value),
                clean(ws.cell(r, c_persona).value),
                clean(ws.cell(r, c_rol).value),
                clean(ws.cell(r, c_otro).value),
                clean(ws.cell(r, c_tel).value),
                clean(ws.cell(r, c_email).value),
                clean(ws.cell(r, c_linkedin).value),
                clean(ws.cell(r, c_comentarios).value),
                clean(ws.cell(r, c_prioritario).value),
            )
        )

    ws = wb["Tipo_Fondo"]
    idx = header_index(ws)
    c_company = resolve_col(idx, ["Company_ID", "CompanyID"], "Tipo_Fondo")
    c_tipo = resolve_col(idx, ["Tipo_Fondo", "Tipo Fondo"], "Tipo_Fondo")
    c_exc = resolve_col(idx, ["Excepciones"], "Tipo_Fondo")
    seen_tipo = set()
    for r in range(2, ws.max_row + 1):
        company_id = parse_int(ws.cell(r, c_company).value)
        tipo = clean(ws.cell(r, c_tipo).value)
        if company_id is None or not tipo:
            continue
        key = (company_id, tipo)
        if key in seen_tipo:
            continue
        seen_tipo.add(key)
        tipo_rows.append((company_id, tipo, clean(ws.cell(r, c_exc).value)))

    ws = wb["Sector"]
    idx = header_index(ws)
    c_company = resolve_col(idx, ["Company_ID", "CompanyID"], "Sector")
    c_sector = resolve_col(idx, ["Sector"], "Sector")
    c_cons = resolve_col(idx, ["Sector_Consolidado"], "Sector")
    seen_sector = set()
    for r in range(2, ws.max_row + 1):
        company_id = parse_int(ws.cell(r, c_company).value)
        sector = clean(ws.cell(r, c_sector).value)
        if company_id is None or not sector:
            continue
        key = (company_id, sector)
        if key in seen_sector:
            continue
        seen_sector.add(key)
        sector_rows.append((company_id, sector, clean(ws.cell(r, c_cons).value)))

    ws = wb["Mercados_Area_Geografica"]
    idx = header_index(ws)
    c_company = resolve_col(idx, ["Company_ID", "CompanyID"], "Mercados_Area_Geografica")
    c_area = resolve_col(idx, ["Mercados_Area_Geografica"], "Mercados_Area_Geografica")
    seen_area = set()
    for r in range(2, ws.max_row + 1):
        company_id = parse_int(ws.cell(r, c_company).value)
        area = clean(ws.cell(r, c_area).value)
        if company_id is None or not area:
            continue
        key = (company_id, area)
        if key in seen_area:
            continue
        seen_area.add(key)
        mapa_rows.append((company_id, area))

    return {
        "inversion": inversion_rows,
        "contactos": contactos_rows,
        "tipo_fondo": tipo_rows,
        "sector": sector_rows,
        "mapa_area_geografica": mapa_rows,
    }


def build_sql(rows: Dict[str, List[Tuple]], do_truncate: bool) -> str:
    lines: List[str] = []
    lines.append("begin;")
    if do_truncate:
        lines.extend(
            [
                "truncate table",
                "  sourcecrm.mapa_area_geografica,",
                "  sourcecrm.sector,",
                "  sourcecrm.tipo_fondo,",
                "  sourcecrm.contactos,",
                "  sourcecrm.inversion",
                "restart identity cascade;",
                "",
            ]
        )

    for row in rows["inversion"]:
        lines.append(
            "insert into sourcecrm.inversion "
            "(company_id, vertical, compania, direccion, estrategia, linkedin, web, portfolio, comentarios, encaje_summax, motivo, inversion_minima, inversion_maxima, prioridad, sede, tamano_empresa) "
            f"values ({row[0]}, {sql_quote(row[1])}, {sql_quote(row[2])}, {sql_quote(row[3])}, {sql_quote(row[4])}, {sql_quote(row[5])}, {sql_quote(row[6])}, {sql_quote(row[7])}, {sql_quote(row[8])}, {sql_quote(row[9])}, {sql_quote(row[10])}, {sql_quote(row[11])}, {sql_quote(row[12])}, {sql_quote(row[13])}, {sql_quote(row[14])}, {sql_quote(row[15])}) "
            "on conflict (company_id) do update set "
            "vertical = excluded.vertical, compania = excluded.compania, direccion = excluded.direccion, estrategia = excluded.estrategia, linkedin = excluded.linkedin, web = excluded.web, portfolio = excluded.portfolio, comentarios = excluded.comentarios, encaje_summax = excluded.encaje_summax, motivo = excluded.motivo, inversion_minima = excluded.inversion_minima, inversion_maxima = excluded.inversion_maxima, prioridad = excluded.prioridad, sede = excluded.sede, tamano_empresa = excluded.tamano_empresa, updated_at = now();"
        )

    for row in rows["contactos"]:
        lines.append(
            "insert into sourcecrm.contactos "
            "(contact_id, company_id, compania, persona_contacto, rol, otro_contacto, telefono, email, linkedin, comentarios, prioritario) "
            f"values ({row[0]}, {row[1]}, {sql_quote(row[2])}, {sql_quote(row[3])}, {sql_quote(row[4])}, {sql_quote(row[5])}, {sql_quote(row[6])}, {sql_quote(row[7])}, {sql_quote(row[8])}, {sql_quote(row[9])}, {sql_quote(row[10])}) "
            "on conflict (contact_id) do update set "
            "company_id = excluded.company_id, compania = excluded.compania, persona_contacto = excluded.persona_contacto, rol = excluded.rol, otro_contacto = excluded.otro_contacto, telefono = excluded.telefono, email = excluded.email, linkedin = excluded.linkedin, comentarios = excluded.comentarios, prioritario = excluded.prioritario, updated_at = now();"
        )

    for row in rows["tipo_fondo"]:
        lines.append(
            "insert into sourcecrm.tipo_fondo (company_id, tipo_fondo, excepciones) "
            f"values ({row[0]}, {sql_quote(row[1])}, {sql_quote(row[2])}) "
            "on conflict (company_id, tipo_fondo) do update set excepciones = excluded.excepciones;"
        )

    for row in rows["sector"]:
        lines.append(
            "insert into sourcecrm.sector (company_id, sector, sector_consolidado) "
            f"values ({row[0]}, {sql_quote(row[1])}, {sql_quote(row[2])}) "
            "on conflict (company_id, sector) do update set sector_consolidado = excluded.sector_consolidado;"
        )

    for row in rows["mapa_area_geografica"]:
        lines.append(
            "insert into sourcecrm.mapa_area_geografica (company_id, area_geografica) "
            f"values ({row[0]}, {sql_quote(row[1])}) "
            "on conflict (company_id, area_geografica) do nothing;"
        )

    lines.extend(
        [
            "",
            "-- Integrity checks (must return 0 in all rows).",
            "select 'contactos_orphans' as check_name, count(*) as orphan_count from sourcecrm.contactos c left join sourcecrm.inversion i on i.company_id = c.company_id where i.company_id is null;",
            "select 'tipo_fondo_orphans' as check_name, count(*) as orphan_count from sourcecrm.tipo_fondo t left join sourcecrm.inversion i on i.company_id = t.company_id where i.company_id is null;",
            "select 'sector_orphans' as check_name, count(*) as orphan_count from sourcecrm.sector s left join sourcecrm.inversion i on i.company_id = s.company_id where i.company_id is null;",
            "select 'mapa_orphans' as check_name, count(*) as orphan_count from sourcecrm.mapa_area_geografica m left join sourcecrm.inversion i on i.company_id = m.company_id where i.company_id is null;",
            "commit;",
        ]
    )

    return "\n".join(lines) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate SQL upserts for sourcecrm from SUMMAX Excel workbook.")
    parser.add_argument("workbook", type=Path, help="Path to workbook (.xlsx)")
    parser.add_argument("--output", type=Path, default=Path("supabase/scripts/load_sourcecrm_from_excel.sql"))
    parser.add_argument("--truncate-first", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook = args.workbook.expanduser().resolve()
    if not workbook.exists():
        raise SystemExit(f"Workbook not found: {workbook}")

    rows = read_excel_rows(workbook)
    sql = build_sql(rows=rows, do_truncate=args.truncate_first)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(sql, encoding="utf-8")

    print(f"Workbook: {workbook.name}")
    print(
        "Rows parsed: "
        f"inversion={len(rows['inversion'])}, "
        f"contactos={len(rows['contactos'])}, "
        f"tipo_fondo={len(rows['tipo_fondo'])}, "
        f"sector={len(rows['sector'])}, "
        f"mapa_area_geografica={len(rows['mapa_area_geografica'])}"
    )
    print(f"SQL generated: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
